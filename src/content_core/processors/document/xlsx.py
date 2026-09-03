import asyncio
from functools import partial

from openpyxl import load_workbook  # type: ignore

from content_core.logging import logger


async def extract_xlsx_content(file_path, max_rows=10000, max_cols=100):
    """Extract content from every worksheet in an XLSX file."""

    def _extract():
        try:
            wb = load_workbook(file_path, data_only=True)
            content = []

            # Prefer worksheets over chartsheets / macrosheets.
            sheets = list(getattr(wb, "worksheets", None) or [])
            if not sheets:
                sheets = [wb[name] for name in wb.sheetnames]

            for ws in sheets:
                title = getattr(ws, "title", None) or "Sheet"
                content.append(f"\n# Sheet: {title}\n")

                max_row = min(ws.max_row or 0, max_rows)
                max_col = min(ws.max_column or 0, max_cols)
                if max_row < 1 or max_col < 1:
                    content.append("_(empty)_")
                    continue

                headers = []
                for col in range(1, max_col + 1):
                    cell_value = ws.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else "")

                content.append("| " + " | ".join(headers) + " |")
                content.append("| " + " | ".join(["---"] * len(headers)) + " |")

                for row in range(2, max_row + 1):
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(
                            str(cell_value) if cell_value is not None else ""
                        )
                    content.append("| " + " | ".join(row_data) + " |")

            wb.close()
            return "\n".join(content)

        except Exception as e:
            logger.error(f"Failed to extract XLSX content: {e}")
            return None

    return await asyncio.get_event_loop().run_in_executor(None, partial(_extract))


async def get_xlsx_info(file_path):
    """Get XLSX metadata and content"""

    async def _get_info():
        try:
            wb = load_workbook(file_path, data_only=True)

            props = {
                "sheet_count": len(wb.sheetnames),
                "sheets": wb.sheetnames,
                "title": wb.properties.title,
                "creator": wb.properties.creator,
                "created": wb.properties.created,
                "modified": wb.properties.modified,
            }

            content = await extract_xlsx_content(file_path)

            stats = {
                "sheet_count": len(wb.sheetnames),
                "total_rows": sum((sheet.max_row or 0) for sheet in wb.worksheets),
                "total_columns": sum((sheet.max_column or 0) for sheet in wb.worksheets),
            }
            wb.close()

            return {"metadata": props, "content": content, "statistics": stats}

        except Exception as e:
            logger.error(f"Failed to get XLSX info: {e}")
            return None

    return await _get_info()
