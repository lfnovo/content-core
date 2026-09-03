"""Multi-sheet XLSX must use the office extractor, not Docling."""

import asyncio

from content_core.config import ContentCoreConfig
from content_core.extraction import _route_for_mime
from content_core.processors.document.docling import DOCLING_SUPPORTED

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def test_xlsx_not_in_docling_supported():
    assert XLSX not in DOCLING_SUPPORTED


def test_auto_routes_xlsx_to_office(monkeypatch):
    import content_core.extraction as extraction

    monkeypatch.setattr(extraction, "DOCLING_AVAILABLE", True)
    monkeypatch.setattr(extraction, "DOCLING_SUPPORTED", DOCLING_SUPPORTED)
    cfg = ContentCoreConfig(document_engine="auto")
    assert _route_for_mime(XLSX, cfg) == "office"


def test_extract_xlsx_includes_all_sheets(tmp_path):
    from openpyxl import Workbook

    from content_core.processors.document.xlsx import extract_xlsx_content

    path = tmp_path / "many.xlsx"
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "h1"
    wb.active["A2"] = "a"
    for i in range(2, 11):
        ws = wb.create_sheet(f"Sheet{i}")
        ws["A1"] = f"h{i}"
        ws["A2"] = f"v{i}"
    wb.save(path)

    content = asyncio.run(extract_xlsx_content(str(path)))
    assert content is not None
    for i in range(1, 11):
        assert f"# Sheet: Sheet{i}" in content
